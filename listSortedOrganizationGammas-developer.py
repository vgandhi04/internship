import os
import math
import csv
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from botocore.exceptions import ClientError
import boto3
from opensearchpy import OpenSearch, RequestsHttpConnection, AWSV4SignerAuth
from boto3.dynamodb.conditions import Key
import AppSyncHelper
import redis
import botocore
import DynamoDBHelper
import json
from datetime import datetime, timedelta
import uuid
import time
import threading
from collections import defaultdict
from chaliceHelper import app
from chalice import BadRequestError, UnauthorizedError, AuthResponse, ChaliceViewError
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
import traceback
from gql_queries import *

OPENSEARCH_DOMAIN_ENDPOINT = os.environ['OPENSEARCH_DOMAIN_ENDPOINT']
OPENSEARCH_PORT = 443
REGION = os.environ['REGION']  # e.g. us-west-1
CREDENTIALS = boto3.Session().get_credentials()
AWS_AUTH = AWSV4SignerAuth(CREDENTIALS, REGION)
STAGE_TABLE_NAME = os.environ['API_STRALIGN_STAGETABLE_NAME']
GAMMA_TABLE_NAME = os.environ['API_STRALIGN_GAMMATABLE_NAME']
# DEPARTMENTGAMMA_TABLE_NAME = os.environ['API_STRALIGN_DEPARTMENTGAMMATABLE_NAME']
DEPARTMENT_TABLE_NAME = os.environ['API_STRALIGN_DEPARTMENTTABLE_NAME']
API_STRALIGN_USERTABLE_NAME = os.environ['API_STRALIGN_USERTABLE_NAME']
API_STRALIGN_PRIORITYBATCHTABLE_NAME = os.environ['API_STRALIGN_PRIORITYBATCHTABLE_NAME']


def get_graphql_client():
    gql_client = AppSyncHelper.create_graphql_client()
    return gql_client


SENDER_EMAIL = "noreply@stralign.com"
GQL_CLIENT = get_graphql_client()
REDIS_CLIENT = redis.Redis(
    host=os.environ["ELASTICACHE_CLUSTER_ENDPOINT"], port=os.environ["ELASTICACHE_CLUSTER_PORT"], db=0)
BUCKET_NAME = os.environ['STORAGE_STRALIGNUPLOADS_BUCKETNAME']
LAMBDA_CLIENT = boto3.client('lambda')
FUNCTION_SCHEDULEDCOMPARISONQUEUER_NAME = os.environ['FUNCTION_SCHEDULEDCOMPARISONQUEUER_NAME']

# @app.authorizer(ttl_seconds=30)
# def my_auth(auth_request):
#     # Validate auth_request.token, and then:
#     print("my_auth called")
#     print("auth_request", auth_request)
#     print("auth_request.token", auth_request.token)
#     return AuthResponse(routes=['/export/rankings'], principal_id='username')


def get_global_ses_client():
    global SES_CLIENT
    if 'SES_CLIENT' not in globals():
        SES_CLIENT = boto3.client('ses')
    return SES_CLIENT


def get_global_opensearch_client():
    global OPENSEARCH_CLIENT
    if 'OPENSEARCH_CLIENT' not in globals():
        OPENSEARCH_CLIENT = OpenSearch(
            hosts=[f'{OPENSEARCH_DOMAIN_ENDPOINT}:{OPENSEARCH_PORT}'],
            http_auth=AWS_AUTH,
            use_ssl=True,
            verify_certs=True,
            connection_class=RequestsHttpConnection
        )
    return OPENSEARCH_CLIENT


def get_global_s3_client():
    global S3_CLIENT
    if 'S3_CLIENT' not in globals():
        S3_CLIENT = boto3.client('s3')
    return S3_CLIENT


def get_global_dynamodb_resource():
    global DYNAMODB_RESOURCE
    if 'DYNAMODB_RESOURCE' not in globals():
        DYNAMODB_RESOURCE = boto3.resource(
            'dynamodb', config=botocore.client.Config(max_pool_connections=1000))
    return DYNAMODB_RESOURCE


def get_global_stage_table():
    global STAGE_TABLE
    if 'STAGE_TABLE' not in globals():
        STAGE_TABLE = get_global_dynamodb_resource().Table(STAGE_TABLE_NAME)
    return STAGE_TABLE


def get_global_department_table():
    global DEPARTMENT_TABLE
    if 'DEPARTMENT_TABLE' not in globals():
        DEPARTMENT_TABLE = get_global_dynamodb_resource().Table(DEPARTMENT_TABLE_NAME)
    return DEPARTMENT_TABLE


# def get_global_department_gamma_table():
#     global DEPARTMENTGAMMA_TABLE
#     if 'DEPARTMENTGAMMA_TABLE' not in globals():
#         DEPARTMENTGAMMA_TABLE = get_global_dynamodb_resource().Table(
#             DEPARTMENTGAMMA_TABLE_NAME)
#     return DEPARTMENTGAMMA_TABLE


def get_global_gamma_table():
    global GAMMA_TABLE
    if 'GAMMA_TABLE' not in globals():
        GAMMA_TABLE = get_global_dynamodb_resource().Table(GAMMA_TABLE_NAME)
    return GAMMA_TABLE


def get_GQL_paginated(gqlClient, query, params, list_query_name, result):
    nextToken = None
    count = 1

    while (nextToken != None or count < 2):
        count += 1

        response = gqlClient.execute(query, variable_values=params)
        # print(response)
        result.extend(response[list_query_name]["items"])
        nextToken = response[list_query_name].get("nextToken", None)
        params["nextToken"] = nextToken
        # print(nextToken)
    return result


def upload_to_s3(file_path, file_name):
    start = time.time()
    get_global_s3_client().upload_file(
        Filename=file_path,
        Bucket=os.environ["STORAGE_STRALIGNUPLOADS_BUCKETNAME"],
        Key="public/" + file_name
    )
    print("time taken to upload file", file_name, time.time()-start)
    return "public/" + file_name


def update_rankings(lambda_client, organization_id, aggregates=False):
    params = {
        "organization_id": organization_id
    }
    if aggregates:
        params['aggregates'] = aggregates
    lambda_client.invoke(
        FunctionName=os.environ['FUNCTION_SCORECALCULATIONSTATUSTRIGGERPYTHON_NAME'],
        InvocationType='Event',
        Payload=json.dumps(params).encode('utf-8'),
    )
    print("invoked rankings updater")
    return


def get_last_dates(today):
    days_to_sunday = (today.weekday() - 6) % 7
    last_sunday_date = today - timedelta(days=days_to_sunday)

    first_day_of_current_month = today.replace(day=1)
    last_month_last_date = first_day_of_current_month - timedelta(days=1)

    quarter_offset = (today.month - 1) % 3 + 1
    first_day_of_current_quarter = today.replace(
        month=today.month - quarter_offset + 1, day=1)
    last_quarter_last_date = first_day_of_current_quarter - timedelta(days=1)

    return [last_sunday_date.strftime('%Y-%m-%d'), last_month_last_date.strftime('%Y-%m-%d'), last_quarter_last_date.strftime('%Y-%m-%d')]


def get_last_login_date(userID, today_date, result):
    query_params = {
        "userID": userID,
        "date": today_date
    }
    last_login_items = get_graphql_client().execute(get_last_login_date_query,
                                                    variable_values=query_params)["loginHistoriesByUserIDAndDate"]["items"]
    if last_login_items:
        result.append(last_login_items[0]['date'])
    return result


def get_s3_rank_history(s3_client, yearly_rank_history, yearly_file_key):
    try:
        s3_response = json.loads(s3_client.get_object(Bucket=BUCKET_NAME, Key=yearly_file_key)[
            'Body'].read().decode('utf-8'))
        # print("s3 response", s3_response)
        yearly_rank_history.update(s3_response)
    except Exception as e:
        print("error", e)
    return yearly_rank_history


# def get_s3_yearly_previous_ranks(s3_client, yearly_rank_history, yearly_file_key, userID, s3_response={}):
#     start = time.time()
#     today_date = datetime.now().date()
#     threads = []
#     result = []
#     threads.append(threading.Thread(target=get_last_login_date,
#                                     args=(userID, today_date.strftime('%Y-%m-%d'), result,)))
#     threads[-1].start()

#     get_s3_rank_history(s3_client, s3_response, yearly_file_key)

#     last_saturday_date, last_month_last_date, last_quarter_last_date = get_last_dates(
#         today_date)
#     for thread in threads:
#         thread.join()
#     last_login_date = result[0] if result else '00-00-0000'
#     yearly_rank_history.update({k: {"Week": v.get(last_saturday_date, {"Rank": -1})["Rank"], "Month": v.get(last_month_last_date, {"Rank": -1})["Rank"],
#                                     "Quarter": v.get(last_quarter_last_date, {"Rank": -1})["Rank"], "lastLogin": v.get(last_login_date, {"Rank": -1})["Rank"]}
#                                for k, v in s3_response.items()})
#     print("time taken to fetch s3 rank history", time.time() - start)
#     return s3_response, yearly_rank_history


def apply_filters(merged_rankings, filters):
    if not filters:
        return merged_rankings

    print("filters", filters)
    owner_filters = filters.get("byMe", "")
    rank_filters = filters.get("Ranks", {})
    stage_filters = filters.get("Stage", {})
    department_filters = filters.get("Department", [])
    created_date_filters = filters.get("Created", [])

    fixed_ranks = bool(rank_filters.get("Assigned", True))
    default_ranks = bool(rank_filters.get("Default", True))
    levels = {stage for stage, val in stage_filters.items() if val}
    departments = set(department_filters)
    user_id = owner_filters
    date_range = [datetime.strptime(
        created_date, '%Y-%m-%dT%H:%M:%S.%fZ') for created_date in created_date_filters]

    filtered_rankings = [
        {
            "Rank": i + 1,
            **ranking
        }
        for i, ranking in enumerate(merged_rankings)
        if (fixed_ranks or ranking["Gamma"]["fixedRank"] <= 0) and
           (default_ranks or ranking["Gamma"]["fixedRank"] >= 0) and
           (not departments or any(item in departments for item in ranking["Gamma"]["departments"])) and
           (not levels or ranking["Gamma"]["level"]["id"] in levels) and
           (not user_id or ranking["Gamma"]["user"]["id"] == user_id) and
           (not date_range or date_range[0] <= datetime.strptime(
               ranking["Gamma"]["createdAt"], '%Y-%m-%dT%H:%M:%S.%fZ') <= date_range[1])
    ]
    return filtered_rankings


def get_default_rankings(gql_client, organizationID, pages, all_rankings, default_rankings):
    start = time.time()
    params = {
        "organizationID": organizationID,
        "nextToken": None
    }

    if pages:

        for page in pages:
            start = (page-1)*50
            end = start + 50
            if start < len(all_rankings):
                keys = [key for key in all_rankings[start: end]]
            else:
                return []

            filter_string = '{{or: [{}]}}'.format(
                ', '.join('{{id: {{eq: "{}"}}}}'.format(key) for key in keys)
            )
            get_gammas_query_modified = get_gammas_query.format(filter_string)
            get_GQL_paginated(
                gql_client, gql(get_gammas_query_modified), params, "gammasByOrganizationID", default_rankings["items"])
    else:
        get_gammas_query_modified = get_gammas_query.format('{}')
        get_GQL_paginated(
            gql_client, gql(get_gammas_query_modified), params, "gammasByOrganizationID", default_rankings["items"])

    print("time taken to fetch default ranking", time.time() - start)
    return default_rankings


def get_all_item_all_ranks(REDIS_CLIENT, priority_batch_id, current_rank=True, previous_rank=True, default_rank=True, no_rank=True):
    items_with_ranks = {}
    if current_rank:
        # Get all items with their ranks from the sorted set
        items_with_ranks = REDIS_CLIENT.zrange(
            priority_batch_id, 0, -1, withscores=True)

    # print("items_with_ranks", items_with_ranks)

    # Get the previous ranks from the hash
    previous_ranks = {}
    if previous_rank:
        previous_ranks = REDIS_CLIENT.zrange(
            f'{priority_batch_id}:previous_rank', 0, -1, withscores=True)
        previous_ranks = {item.decode('utf-8'): int(rank)
                          for item, rank in previous_ranks}

    # print("previous_ranks", previous_ranks)
    no_ranks = {}
    if no_rank:
        no_ranks = REDIS_CLIENT.zrange(
            f'{priority_batch_id}:no_rank', 0, -1, withscores=True)

    # Get the default ranks from the hash
    default_ranks = {}
    if default_rank:
        default_ranks = REDIS_CLIENT.hgetall(
            f'{priority_batch_id}:default_rank')

    # print("default_ranks", default_ranks)

    # Combine the ranks and previous ranks into a dictionary
    redis_ranks = {
        item.decode('utf-8'): {
            'rank': int(rank),
            'previous_rank': int(previous_ranks.get(item, -1)),
            'default_rank': int(default_ranks.get(item, -1))
        }
        for item, rank in items_with_ranks
    }

    no_ranks = {
        item.decode('utf-8'): createdAt
        for item, createdAt in no_ranks
        if not redis_ranks.get(item.decode('utf-8'))
    }
    # Sort the dictionary by rank in ascending order
    redis_ranks = dict(sorted(redis_ranks.items(), key=lambda x: x[1]['rank']))

    return redis_ranks, no_ranks


def get_redis_rankings(REDIS_CLIENT, priorityBatchID, redis_rankings, no_rankings, current_rank=True, previous_rank=True, default_rank=True, no_rank=True):
    start = time.time()
    redis_rankings["items"], no_rankings["items"] = get_all_item_all_ranks(
        REDIS_CLIENT, priorityBatchID, current_rank, previous_rank, default_rank, no_rank)
    # print("redis_rankings", redis_rankings)
    print("time taken to fetch redis ranking", time.time()-start)
    return redis_rankings, no_rankings


def get_redis_rank(REDIS_CLIENT, priorityBatchID, itemID, rank):
    rank.append(REDIS_CLIENT.zscore(priorityBatchID, itemID))
    return rank


def create_ranking_excel_file(merged_rankings, timenow, file_keys):
    HEADER_ROW_HEIGHT = 20
    ROW_HEIGHT = 17
    header_row_fill = PatternFill(
        start_color='DFEDF9', end_color='DFEDF9', fill_type="solid")
    data_row_even_fill = PatternFill(
        start_color='F7FAFD', end_color='F7FAFD', fill_type="solid")
    data_row_odd_fill = PatternFill(
        start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
    # Create a new workbook and select the active sheet
    workbook = Workbook()
    sheet = workbook.active

    headers = ['Rank', 'Title', 'Stage', 'Change', 'Previous']
    header_row_border = Border(
        bottom=Side(style='thin'))
    # adjust header row height for better readability
    sheet.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Apply formatting to the header row
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.alignment = Alignment(horizontal='center')
        # cell.font = cell.font.copy(size=12, bold=True)
        cell.font = Font(size=12, bold=True)
        cell.fill = header_row_fill
        cell.border = header_row_border

    # Iterate over the merged rankings and populate the spreadsheet
    for item in merged_rankings:
        # rank = i + 1 if item['isRanked'] else '-'
        rank = item['Gamma']['fixedRank'] if item['Gamma']['fixedRank'] > 0 else (
            item['Rank'] if item['isRanked'] else '-')
        title = item['Gamma']['title']
        stage = item['Gamma']['level']['name']
        if item['Gamma']['fixedRank'] < 0:

            previous = '-' if (item['Gamma'].get('previousRank', None)
                               ) in [None, -1] else item['Gamma']['previousRank']
        else:
            previous = item['Gamma']['fixedRank']
        change = previous - \
            rank if isinstance(previous, int) and isinstance(
                rank, int) else '-'

        sheet.append([rank, title, stage, change, previous])

    column_dimensions = [
        ('A', 10), ('B', 80), ('C', 18), ('D', 10), ('E', 10)
    ]
    for col, width in column_dimensions:
        sheet.column_dimensions[col].width = width

    # Apply formatting to data rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        i = row[0].row
        # adjust data row height for better readability
        sheet.row_dimensions[i].height = ROW_HEIGHT
        row_fill = data_row_even_fill if i % 2 == 0 else data_row_odd_fill

        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal='center')
            # cell.font = cell.font.copy(size=12)
            cell.font = Font(size=12, bold=False)
            cell.fill = row_fill

    # Save the workbook
    file_name = "Rankings" + timenow + ".xlsx"
    file_path = "/tmp/" + file_name
    workbook.save(file_path)
    # file_key = upload_to_s3(file_path, file_name)
    file_keys.append(file_path)
    return file_path


def create_excel_file(merged_rankings, users, objectives, departments, vote_counts, rating_counts, export_data):
    file_keys = []
    threads = []
    timenow = datetime.now().strftime("%m-%d-%Y%H-%M-%S.%f")

    threads.append(threading.Thread(target=create_ranking_excel_file,
                                    args=(merged_rankings, timenow, file_keys,)))
    threads[-1].start()
    if export_data == 'All':
        vote_sheet_headers = ['Title', 'User']
        vote_sheet_headers.extend(
            [f"{objective['name']} ({status})" for objective in objectives for status in ['Compared', 'Selected']])

        rating_sheet_headers = ['Title', 'User']
        rating_sheet_headers.extend(
            [f"{objective['name']} ({status})" for objective in objectives for status in ['Rating Sum']])

        # Writing vote_counts to CSV
        vote_counts_file_name = "/tmp/Votes" + timenow + ".csv"
        file_keys.append(vote_counts_file_name)
        with open(vote_counts_file_name, 'w', newline='') as csvfile:

            writer = csv.DictWriter(csvfile, fieldnames=vote_sheet_headers)

            # Write the header rows
            # First header row
            writer.writerow(dict(zip(vote_sheet_headers, vote_sheet_headers)))

            # Write the data
            for item in merged_rankings:

                gamma_value = vote_counts[item['Gamma']['id']]
                for user in users:
                    row = {"Title": item['Gamma']
                           ['title'], "User": user['email']}

                    for objective in objectives:
                        objective_value = gamma_value[objective['id']]
                        user_value = objective_value[user['id']]
                        row.update(
                            {f"{objective['name']} (Compared)": user_value['comparison_count'],
                             f"{objective['name']} (Selected)": user_value['selection_count']})

                    writer.writerow(row)

        # Writing rating_counts to CSV
        rating_counts_file_name = "/tmp/Ratings" + timenow + ".csv"
        file_keys.append(rating_counts_file_name)
        with open(rating_counts_file_name, 'w', newline='') as csvfile:

            writer = csv.DictWriter(csvfile, fieldnames=rating_sheet_headers)

            # Write the header rows
            # First header row
            writer.writerow(
                dict(zip(rating_sheet_headers, rating_sheet_headers)))

            # Write the data
            for item in merged_rankings:

                gamma_value = vote_counts[item['Gamma']['id']]
                for user in users:
                    row = {"Title": item['Gamma']
                           ['title'], "User": user['email']}

                    for objective in objectives:
                        objective_value = gamma_value[objective['id']]
                        user_value = objective_value[user['id']]
                        row.update(
                            {f"{objective['name']} (Rating Sum)": user_value})

                    writer.writerow(row)

    for thread in threads:
        thread.join()
    return file_keys


def get_departments_by_organization(gql_client, organization_id, departments):
    params = {
        'organizationID': organization_id
    }

    res = gql_client.execute(
        get_departments_by_organization_query, variable_values=params).get('departmentsByOrganizationID', {"items": []})['items']
    departments.update(
        {department['id']: department['name'] for department in res})
    return departments


def get_user_emails(user_ids, user_details):
    user_keys = [{'id': user_id} for user_id in user_ids]

    users = DynamoDBHelper.batch_get(get_global_dynamodb_resource(), API_STRALIGN_USERTABLE_NAME, user_keys, [
        'id', 'email', 'firstName', 'lastName'])
    user_details.update({user['id']: user for user in users})
    return user_details

# may need to modify dynamodb helper batch get method to use ExpressionAttributeNames


def get_department_names(department_ids, department_details):
    department_keys = [{'id': department_id}
                       for department_id in department_ids]
    expression_attribute_names = {'#n': 'name'}

    departments = DynamoDBHelper.batch_get(get_global_dynamodb_resource(), DEPARTMENT_TABLE_NAME, department_keys, [
        'id', '#n'], expression_attribute_names)
    department_details.update(
        {department['id']: department['name'] for department in departments})
    return department_details


def get_organization_user_vote_aggregates(gql_client, organization_id, gamma_vote_aggregates):
    params = {
        'organizationID': organization_id
    }

    gamma_vote_aggregates.extend(gql_client.execute(
        get_organization_user_vote_aggregates_query, variable_values=params).get('getOrganizationGammaUserVotesAggregates', {'buckets': []})['buckets'])
    # gamma_vote_aggregates.update(
    #     {gamma_bucket['key']: gamma_bucket for gamma_bucket in res['buckets']})
    return gamma_vote_aggregates


def get_organization_user_rating_aggregates(gql_client, organization_id, gamma_rating_aggregates):
    params = {
        'organizationID': organization_id
    }

    gamma_rating_aggregates.extend(gql_client.execute(
        get_organization_user_rating_aggregates_query, variable_values=params).get('getOrganizationGammaUserRatingsAggregates', {'buckets': []})['buckets'])
    # gamma_rating_aggregates.update(
    #     {gamma_bucket['key']: gamma_bucket for gamma_bucket in res['buckets']})
    return gamma_rating_aggregates


def get_users_by_organization(gql_client, organization_id, users):
    params = {
        'organizationID': organization_id,
        'nextToken': None
    }
    users = get_GQL_paginated(
        gql_client, get_users_by_organization_query, params, "usersByOrganizationID", users)

    return users


def get_objectives_by_organization(gql_client, organization_id, objectives):
    params = {
        'organizationObjectivesId': organization_id,
        'nextToken': None
    }
    objectives = get_GQL_paginated(
        gql_client, get_objectives_by_organization_query, params, "objectivesByOrganizationObjectivesId", objectives)

    return objectives


def send_email_attachment(destination, subject, username, file_keys):
    # emailData = {
    #     "email": "email",
    #     "password": "password",
    #     "login_link": "https://d25d5goqjgy5td.cloudfront.net/login"
    # }
    file_names = [key.split('/')[-1] for key in file_keys]
    file_name = "<br>".join(file_names)

    emailData = {
        "username": username,
        "filename": file_name,
    }
    now = datetime.now()
    est = now - timedelta(hours=4)
    try:
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = destination

        msg['Subject'] = est.strftime('%m-%d-%Y') + " " + subject
        body = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8" />
            <meta http-equiv="X-UA-Compatible" content="IE=edge" />
            <meta name="viewport" content="width=device-width, initial-scale=1.0" />
            <title>StrAlign</title>

            <style>
            body,
            html {{
                font-family: Arial, Helvetica, sans-serif;
            }}
            p {{
                font-size: 15px;
                line-height: 22px;
                color: #161616;
            }}
            </style>
        </head>
        <body topmargin="0">
            <table
            width="600"
            border="0"
            align="center"
            cellpadding="0"
            cellspacing="0"
            class="bgtext"
            >
            <thead>
                <tr>
                <th width="100%;" style="display: block; margin: 20px 0">
                    <a href="#"
                    ><img
                        class="header-logo-image"
                        src="https://stralign-static-files.s3.amazonaws.com/logo.png"
                        alt="StrAlign"
                        title="StrAlign"
                    /></a>
                </th>
                </tr>
            </thead>

            <tbody style="background-color: #f8f8f8">
                <tr>
                <td colspan="1">
                    <table
                    width="100%"
                    border="0"
                    align="center"
                    cellpadding="8"
                    cellspacing="0"
                    class="bgtext"
                    >
                    <tr>
                        <div
                        style="
                            background-color: white;
                            border-radius: 12px;
                            width: 85%;
                            margin: 20px auto 20px;
                            padding: 5px 20px 30px;
                        "
                        >
                        <h4><b>Dear {username},</b></h4>
                        <p>Please find the exported file attachments below:</p>
                        <p>
                            {filename}
                        </p>
                        </div>
                    </tr>

                    <tr>
                        <th
                        width="100%;"
                        style="display: block; margin: 20px 0 0; padding-bottom: 0"
                        >
                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/Facebook.png"
                            alt="fb"
                            style="height: 21px"
                        /></a>
                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/TwitterCircle.png"
                            alt="insta"
                            style="height: 21px"
                        /></a>

                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/Linkedin.png"
                            alt="linkedin"
                            style="height: 21px"
                        /></a>
                        </th>
                        <th
                        width="100%;"
                        style="display: block; margin: 0; padding-top: 0"
                        >
                        <p
                            style="
                            margin-top: 0;
                            color: #646464;
                            font-weight: 500;
                            font-size: 13px;
                            "
                        >
                            © 2024 Echo Consulting, LLC<br />
                            Copyright all rights reserved.
                        </p>
                        </th>
                    </tr>
                    </table>
                </td>
                </tr>
            </tbody>
            <tfoot>
                <tr>
                <td></td>
                <td></td>
                </tr>
            </tfoot>
            </table>
        </body>
        </html>


        """.format(**emailData)
        msg.attach(MIMEText(body, 'html'))

        # add the attachments
        attachments = file_keys
        for attachment in attachments:
            part = MIMEBase('application', "octet-stream")
            with open(attachment, 'rb') as file:
                part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition', f'attachment; filename="{os.path.basename(attachment)}"')
            msg.attach(part)

        # send the email
        raw_msg = msg.as_string()
        response = get_global_ses_client().send_raw_email(
            Source=SENDER_EMAIL,
            Destinations=[destination],
            RawMessage={'Data': raw_msg}
        )
        print("Email sent! Message ID:", response['MessageId'])

    except ClientError as e:
        print("Email not sent. Error message: ",
              e.response['Error']['Message'])

    return {
        'statusCode': 200
    }


@app.route('/export/rankings', cors=True, methods=['POST'],  content_types=['application/json'])
def export_rankings():
    event = app.current_request.to_dict()
    body = app.current_request.json_body
    print('received event:')
    print(event)
    print("body", body)
    # read filters if any
    priority_batch_id = body["priorityBatchID"]
    organization_id = body["id"]
    filters = body.get("filters", {})
    export_data = body.get("exportData", "Rankings")

    threads = []
    users = []
    objectives = []
    redis_rankings = {}
    default_rankings = {}
    departments = {}
    no_rankings = {}
    gamma_vote_aggregates = []
    gamma_rating_aggregates = []

    # default_rankings = {}
    # redis_rankings = {}
    # no_rankings = {}
    # threads = []

    # threads.append(threading.Thread(target=get_redis_rankings,
    #                                 args=(REDIS_CLIENT, priorityBatchID, redis_rankings, no_rankings)))
    # threads[-1].start()

    # for thread in threads:
    #     thread.join()
    # all_rankings = list(
    #     redis_rankings["items"].keys()) + list(no_rankings["items"].keys())
    # all_ranks = len(all_rankings)
    # pages = [i for i in range(1, math.ceil(all_ranks/50) + 1)]

    # threads = []
    # default_rankings["items"] = []

    # threads.append(threading.Thread(target=get_default_rankings,
    #                                 args=(GQL_CLIENT, organizationID, pages, all_rankings, default_rankings,)))

    # threads[-1].start()

    # for thread in threads:
    #     thread.join()

    # print("default_rankings", default_rankings["items"])

    # print("redis_rankings", redis_rankings["items"])
    # default_rankings["items"] = {
    #     item["id"]: item for item in default_rankings["items"]}

    # merged_rankings = []
    # merged_rankings = [{"isRanked": True, "Gamma": {**default_rankings["items"][k], "rank": v["rank"], "previousRank": v["previous_rank"]}}
    #                    for k, v in redis_rankings["items"].items() if k in default_rankings["items"]]

    # merged_rankings += [{"isRanked": False, "Gamma": {**item, "rank": 0, "previousRank": None}}
    #                     for item in default_rankings["items"].values() if item["id"] not in redis_rankings["items"]]

    # print("before filters", merged_rankings)
    # # apply filters here:
    # merged_rankings = apply_filters(merged_rankings, filters)
    # print("after filters", merged_rankings)
    # file_key = create_excel_file(merged_rankings)

    # payload = {"fileKey": file_key}

    default_rankings["items"] = []
    threads.append(threading.Thread(target=get_default_rankings,
                                    args=(GQL_CLIENT, organization_id, [], [], default_rankings,)))

    threads[-1].start()

    threads.append(threading.Thread(target=get_users_by_organization,
                                    args=(get_graphql_client(), organization_id, users,)))
    threads[-1].start()

    threads.append(threading.Thread(target=get_redis_rankings,
                                    args=(REDIS_CLIENT, priority_batch_id, redis_rankings, no_rankings)))
    threads[-1].start()

    threads.append(threading.Thread(target=get_organization_user_vote_aggregates,
                                    args=(get_graphql_client(), organization_id, gamma_vote_aggregates,)))
    threads[-1].start()

    threads.append(threading.Thread(target=get_organization_user_rating_aggregates,
                                    args=(get_graphql_client(), organization_id, gamma_rating_aggregates,)))
    threads[-1].start()

    threads.append(threading.Thread(target=get_departments_by_organization,
                                    args=(get_graphql_client(), organization_id, departments,)))
    threads[-1].start()

    threads.append(threading.Thread(target=get_objectives_by_organization,
                                    args=(get_graphql_client(), organization_id, objectives,)))
    threads[-1].start()

    for thread in threads:
        thread.join()
    print("objectives", objectives)

    vote_counts = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: {'comparison_count': 0, 'selection_count': 0})))
    rating_counts = defaultdict(lambda: defaultdict(lambda: defaultdict(
        lambda: 0)))

    for gamma_bucket in gamma_vote_aggregates:
        for objective_bucket in gamma_bucket['buckets']:
            for vote_bucket in objective_bucket['votes']:
                if int(vote_bucket['key']) == -1:
                    for unique_user in vote_bucket['unique_users']:
                        vote_counts[gamma_bucket['key']][objective_bucket['key']
                                                         ][unique_user['key']]['comparison_count'] += unique_user['doc_count']
                elif int(vote_bucket['key']) == 1:
                    for unique_user in vote_bucket['unique_users']:
                        vote_counts[gamma_bucket['key']][objective_bucket['key']
                                                         ][unique_user['key']]['comparison_count'] += unique_user['doc_count']
                        vote_counts[gamma_bucket['key']][objective_bucket['key']
                                                         ][unique_user['key']]['selection_count'] += unique_user['doc_count']

    for gamma_bucket in gamma_rating_aggregates:
        for objective_bucket in gamma_bucket['buckets']:
            for unique_user in objective_bucket['unique_users']:
                rating_counts[gamma_bucket['key']][objective_bucket['key']
                                                   ][unique_user['key']] += unique_user['ratings_sum']['value']

    print("redis_rankings", redis_rankings)
    print("no_rankings", no_rankings)
    all_rankings = list(
        redis_rankings["items"].keys()) + list(no_rankings["items"].keys())
    all_ranks = len(all_rankings)

    default_rankings["items"] = {
        item["id"]: item for item in default_rankings["items"]}

    merged_rankings = []
    merged_rankings = [{"isRanked": True, "Gamma": {**default_rankings["items"][k], "rank": v["rank"], "previousRank": v["previous_rank"]}}
                       for k, v in redis_rankings["items"].items() if k in default_rankings["items"]]

    merged_rankings += [{"isRanked": False, "Gamma": {**item, "rank": 0, "previousRank": None}}
                        for item in default_rankings["items"].values() if item["id"] not in redis_rankings["items"]]
    print("before filters", merged_rankings)
    # apply filters here:
    merged_rankings = apply_filters(merged_rankings, filters)
    print("after filters", merged_rankings)
    file_keys = create_excel_file(merged_rankings, users, objectives,
                                  departments, vote_counts, rating_counts, export_data)
    if os.environ["ENV"] in ["client", "live"]:
        s3_file_keys = []
        for file_key in file_keys:
            s3_file_keys.append(upload_to_s3(
                file_key, file_key.split('/')[-1]))
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Headers': '*',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'POST'
            },
            'body': {"fileKeys": s3_file_keys}
        }
    else:
        send_email_attachment(
            body['email'], "Rankings Export Summary", body['name'], file_keys)

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'POST'
        },
        'body': "Successfully sent the email"
    }


def get_redis_default_ranks(REDIS_CLIENT, priority_batch_id):
    return REDIS_CLIENT.hgetall(
        f'{priority_batch_id}:default_rank')


def get_redis_ranks(REDIS_CLIENT, priority_batch_id, start_rank, end_rank, sort_direction="ASC"):
    items = REDIS_CLIENT.zrange(
        priority_batch_id, start_rank, end_rank, withscores=True)
    return items


def get_redis_ranks_threaded(REDIS_CLIENT, priority_batch_id, start_rank, end_rank, ranks, sort_direction="ASC"):
    res = REDIS_CLIENT.zrange(
        priority_batch_id, start_rank, end_rank, withscores=True)
    ranks.update({item.decode('utf-8'): int(rank)
                  for item, rank in res})

    return ranks


def get_redis_ranks_count(REDIS_CLIENT, priority_batch_id):
    total_ranks = REDIS_CLIENT.zcard(priority_batch_id)
    return total_ranks


def get_redis_previous_ranks(REDIS_CLIENT, priority_batch_id, start_rank, end_rank, sort_direction="ASC"):
    items = REDIS_CLIENT.zrange(
        f'{priority_batch_id}:previous_rank', start_rank, end_rank, withscores=True)
    return items


def get_redis_previous_ranks_count(REDIS_CLIENT, priority_batch_id):

    total_previous_ranks = REDIS_CLIENT.zcard(
        f'{priority_batch_id}:previous_rank')
    return total_previous_ranks


def get_redis_no_ranks(REDIS_CLIENT, priority_batch_id, start_rank, end_rank, sort_direction="ASC"):
    items = REDIS_CLIENT.zrange(
        f'{priority_batch_id}:no_rank', start_rank, end_rank, withscores=True)
    return items


def get_redis_no_ranks_count(REDIS_CLIENT, priority_batch_id):
    total_no_ranks = REDIS_CLIENT.zcard(
        f'{priority_batch_id}:no_rank')
    return total_no_ranks


def get_redis_no_previous_ranks(REDIS_CLIENT, priority_batch_id, start_rank, end_rank, sort_direction="ASC"):
    items = REDIS_CLIENT.zrange(
        f'{priority_batch_id}:no_previous_rank', start_rank, end_rank, withscores=True)
    return items


def get_redis_no_previous_ranks_count(REDIS_CLIENT, priority_batch_id):
    total_no_previous_ranks = REDIS_CLIENT.zcard(
        f'{priority_batch_id}:no_previous_rank')
    return total_no_previous_ranks


def query(table, index_name, key, select, items):
    try:
        res = table.query(
            IndexName=index_name,
            Select='ALL_ATTRIBUTES' if not select else 'SPECIFIC_ATTRIBUTES',
            ProjectionExpression=",".join(select),
            KeyConditionExpression=Key('organizationID').eq(key)

        )
        if res.get('Items'):
            for item in res['Items']:
                items[item['gammaId']].append(item['departmentId'])
            # items[key].extend([mapping['departmentId']
            #                   for mapping in res['Items']])
    except Exception as e:
        print("error", e)
        return items
    return items


# def get_gamma_department_relations(gammas, departments_gammas, query_threads):
#     department_gamma_table = get_global_department_gamma_table()
#     for gamma in gammas:
#         query_threads.append(threading.Thread(
#             target=query, args=(department_gamma_table, 'byGamma', gamma, ['departmentId', 'gammaId'], departments_gammas,)))
#         query_threads[-1].start()
#     return departments_gammas


# def get_gamma_department_relations_by_organization(organization_id, departments_gammas, query_threads):
#     department_gamma_table = get_global_department_gamma_table()
#     query_threads.append(threading.Thread(
#         target=query, args=(department_gamma_table, 'byOrganization', organization_id, ['departmentId', 'gammaId'], departments_gammas,)))
#     query_threads[-1].start()
#     return departments_gammas


# def apply_department_filters(organization_id, all_items, departments_gammas, filter, query_threads):
#     # get_gamma_department_relations(
#     #     all_items, departments_gammas, query_threads)
#     if query_threads:
#         for thread in query_threads:
#             thread.join()
#     else:
#         get_gamma_department_relations_by_organization(
#             organization_id, departments_gammas, query_threads)
#         for thread in query_threads:
#             thread.join()

#     all_items = [item for item in all_items if any(
#         department in departments_gammas[item] for department in filter['Department'])]
#     return all_items


def get_stages_departments(organization_id, stage_details, department_details):
    stages_response = get_global_stage_table().query(KeyConditionExpression=Key(
        'organizationID').eq(organization_id),
        IndexName='byOrganization',
        ProjectionExpression='id,#n,#l',
        ExpressionAttributeNames={'#n': 'name', '#l': 'level'}
    )

    print("stages_response", stages_response)

    stage_details.update({
        stage['id']: {'id': stage['id'], 'name': stage['name'], 'level': int(stage['level'])} for stage in stages_response['Items']})

    departments_response = get_global_department_table().query(KeyConditionExpression=Key(
        'organizationID').eq(organization_id),
        IndexName='byOrganization',
        ProjectionExpression='id,#n',
        ExpressionAttributeNames={'#n': 'name'}
    )

    print("departments_response", departments_response)

    department_details.update({
        department['id']: department['name'] for department in departments_response['Items']})


def get_priority_batches(priority_batch_keys, priority_batches):
    priority_batch_keys = [{"id": priority_batch}
                           for priority_batch in priority_batch_keys if priority_batch]
    priority_batches.update({priority_batch['id']: priority_batch for priority_batch in DynamoDBHelper.batch_get(get_global_dynamodb_resource(), API_STRALIGN_PRIORITYBATCHTABLE_NAME, priority_batch_keys, [
        'id', 'voteFilters', 'measurementFieldFilters', 'fieldFilters'])})
    return priority_batches


def calculate_merged_rankings(row, response, gamma_measurement_fields_aggregates, priority_batches, priority_batch_id, priority_batch_id1, priority_batch_id2, rank_items, rank_items1, rank_items2, previous_rank_items, default_rank_items, stage_details, department_details):
    merged_rankings = []
    print("priority_batches", priority_batches)

    match = [
        priority_batches[priority_batch_id]['measurementFieldFilters']['match'] if priority_batch_id and priority_batches[priority_batch_id].get(
            'measurementFieldFilters', {}).get('filters') else None,
        priority_batches[priority_batch_id1]['measurementFieldFilters']['match'] if priority_batch_id1 and priority_batches[priority_batch_id1].get(
            'measurementFieldFilters', {}).get('filters') else None,
        priority_batches[priority_batch_id2]['measurementFieldFilters']['match'] if priority_batch_id2 and priority_batches[priority_batch_id2].get(
            'measurementFieldFilters', {}).get('filters') else None
    ]
    total_gammas = len(response['hits']['hits'])

    cumulative_sums = {}
    today_date = datetime.now().date()
    previous_rank_dates = get_last_dates(today_date)

    for i in range(total_gammas):
        previous_item = response['hits']['hits'][i-1] if i > 0 else {}
        # Initialize cumulative sum for each item
        item_cumulative_sum = dict(
            cumulative_sums[previous_item['_source']['id']]) if previous_item else {}
        item = response['hits']['hits'][i]
        for field in item["_source"].get("measurementFields", []):
            field_id = field["measurementFieldID"]
            value = field["value"]
            item_cumulative_sum[field_id] = item_cumulative_sum.get(
                field_id, 0) + value  # Update cumulative sum for the current item
        # Store cumulative sum for the item
        cumulative_sums[item["_id"]] = item_cumulative_sum
    print("cumulative_sums", cumulative_sums)
    for i, gamma in enumerate(response['hits']['hits']):
        outOfThreshold = [False, False, False]
        row += 1
        # gamma_measurement_fields_aggregates.clear()
        for measurement_unit in gamma['_source'].get('measurementFields', []):
            field_id = measurement_unit['measurementFieldID']
            value = measurement_unit['value']
            gamma_measurement_fields_aggregates[field_id] = gamma_measurement_fields_aggregates.get(
                field_id, 0) + value

        for batch_id, outOfThresholdIndex in zip([priority_batch_id, priority_batch_id1, priority_batch_id2], range(3)):
            if batch_id:
                threshold_checks = []
                if not priority_batches[batch_id].get('measurementFieldFilters', {}).get('filters') or outOfThreshold[outOfThresholdIndex]:
                    continue

                filters = priority_batches[batch_id]['measurementFieldFilters']['filters']

                for filter in filters:
                    threshold = filter.get('threshold', None)
                    field_id = filter['measurementFieldID']
                    field = filter['measurementField']

                    if threshold is not None and field_id in gamma_measurement_fields_aggregates:
                        if field['type'] == '+-':

                            if field['subType'] == '+-':

                                if (gamma_measurement_fields_aggregates[field_id]/row) < threshold:
                                    threshold_met = False
                                    for j, next_gamma in enumerate(response['hits']['hits'][i+1:], start=i+1):
                                        if cumulative_sums[next_gamma['_source']['id']][field_id]/(j+1) >= threshold:
                                            threshold_checks.append(False)
                                            threshold_met = True
                                            break
                                    if not threshold_met:
                                        threshold_checks.append(True)
                                else:
                                    threshold_checks.append(False)
                            else:
                                if (gamma_measurement_fields_aggregates[field_id]/row) > threshold:
                                    threshold_met = False
                                    for j, next_gamma in enumerate(response['hits']['hits'][i+1:], start=i+1):
                                        print(j, next_gamma)
                                        if cumulative_sums[next_gamma['_source']['id']][field_id]/(j+1) <= threshold:
                                            threshold_checks.append(False)
                                            threshold_met = True
                                            break
                                    if not threshold_met:
                                        threshold_checks.append(True)

                                else:
                                    threshold_checks.append(False)

                        else:
                            if gamma_measurement_fields_aggregates[field_id] > threshold:
                                threshold_checks.append(True)
                            else:
                                threshold_checks.append(False)

                    else:
                        threshold_checks.append(False)

                if match[outOfThresholdIndex] == 'ONE':
                    outOfThreshold[outOfThresholdIndex] = any(threshold_checks)
                elif match[outOfThresholdIndex] == 'ALL':
                    outOfThreshold[outOfThresholdIndex] = all(threshold_checks)
        previous_rank_history = {}
        if gamma.get('inner_hits', {}):
            for unit in gamma['inner_hits']['rankHistory']['hits']['hits']:

                date = unit['_source']["D"]
                rank = unit['_source']["R"]
                unit['_source']["S"] = stage_details.get(unit['_source']["S"], None)
                #if date in previous_rank_dates
                previous_rank_history[date] = {"R": rank, "S": unit['_source']["S"]}
        merged_rankings.append({
            "isRanked": True if rank_items.get(gamma['_source']["id"]) else False,
            "Gamma": {
                "id": gamma['_source']["id"],
                "friendlyId": gamma['_source']["friendlyId"],
                "title": gamma['_source']["title"],
                "description": gamma['_source']["description"],
                "fixedRank": gamma['_source']["fixedRank"],
                "level": stage_details[gamma['_source']["levelID"]],
                "departments": [{"id": department, "name": department_details[department]} for department in gamma['_source']["departments"] if department_details.get(department)],
                "rank": rank_items.get(gamma['_source']["id"], 0),
                "rank1": rank_items1.get(gamma['_source']["id"], 0),
                "rank2": rank_items2.get(gamma['_source']["id"], 0),
                "outOfThreshold": outOfThreshold[0],
                "outOfThreshold1": outOfThreshold[1],
                "outOfThreshold2": outOfThreshold[2],
                "previousRank": {
                    "Recent": previous_rank_items.get(gamma['_source']["id"], -1),
                    "Week": previous_rank_history.get(previous_rank_dates[0], {"R": -1})['R'],
                    "Month": previous_rank_history.get(previous_rank_dates[1], {"R": -1})['R'],
                    "Quarter": previous_rank_history.get(previous_rank_dates[2],  {"R": -1})['R']
                },
                "defaultRank": default_rank_items.get(gamma['_source']["id"], -1),
                "rankHistory": previous_rank_history
            }
        })

    return merged_rankings


def last_six_sundays():
    today = datetime.now().date()
    weekday = today.weekday()
    # Calculate the date of the last Sunday
    last_sunday = today - timedelta(days=weekday + 1)

    # List to store the dates
    sundays = []
    for _ in range(6):
        sundays.append(last_sunday.strftime("%Y-%m-%d"))
        last_sunday -= timedelta(days=7)  # Move to the previous Sunday

    return sundays


def generate_dates(start_date_str, end_date_str, interval):
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
    dates_list = []

    current_date = start_date
    while current_date <= end_date:
        if interval == 'day':
            dates_list.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        elif interval == 'week':
            if current_date.weekday() == 6:  # Sunday
                dates_list.append(current_date.strftime('%Y-%m-%d'))
            current_date += timedelta(days=1)
        elif interval == 'month':
            # Go to the last day of the month
            next_month = current_date.replace(day=28) + timedelta(days=4)
            last_day_of_month = next_month - timedelta(days=next_month.day)
            dates_list.append(last_day_of_month.strftime('%Y-%m-%d'))
            current_date = last_day_of_month + timedelta(days=1)

    return dates_list


@app.route('/rankings', cors=True, methods=['GET'], content_types=['application/json'])
def get_rankings():
    PAGE_SIZE = 1000
    OPENSEARCH_CLIENT = get_global_opensearch_client()
    INDEX_NAME = 'gamma'

    priority_batch_id = app.current_request.query_params["priorityBatchID"]
    priority_batch_id1 = app.current_request.query_params.get(
        "priorityBatchID1", None)
    priority_batch_id2 = app.current_request.query_params.get(
        "priorityBatchID2", None)

    organization_id = app.current_request.query_params["id"]
    user_id = app.current_request.query_params.get("userID", None)
    page = int(app.current_request.query_params.get("page", 1))
    sort_field = app.current_request.query_params.get("sort_field", "rank")
    search = app.current_request.query_params.get("search", None)
    if sort_field in ["position1", "position2"]:
        sort_field = "rank"
    ranks_only = app.current_request.query_params.get("ranks_only", False)
    sort_direction = app.current_request.query_params.get("sort_direction", "ASC")
    filter = app.current_request.query_params.get('filter', "{}")

    print("filter", filter)
    filter = json.loads(filter)

    chart = app.current_request.query_params.get('chart', "{}")
    print("chart", chart)
    chart = json.loads(chart)

    nextToken = app.current_request.query_params.get("nextToken", "{}")
    nextToken = json.loads(nextToken)
    authorizer = app.current_request.context.get("authorizer", {})
    claims = authorizer.get("claims", {})
    authorized = False
    if claims.get("SUPERADMIN"):
        authorized = True
    elif claims.get("organization"):
        if organization_id == claims['organization']:
            authorized = True

    if not authorized:
        raise UnauthorizedError("Not authorized")

    if page == "null":
        page = None
    payload = {}
    threads = []
    # start = (page-1) * PAGE_SIZE
    start = 0
    # end = page * PAGE_SIZE
    end = 1000

    rank_items = []
    no_rank_items = []

    rank_items1 = {}
    rank_items2 = {}

    priority_batches = {}
    gamma_measurement_fields_aggregates = defaultdict(int, nextToken)
    threads.append(threading.Thread(target=get_priority_batches,
                                    args=(set([priority_batch_id, priority_batch_id1, priority_batch_id2]), priority_batches,)))
    threads[-1].start()

    # sort can be applied on rank/title/level/change/previousRank

    if priority_batch_id1:

        threads.append(threading.Thread(target=get_redis_ranks_threaded,
                                        args=(REDIS_CLIENT, priority_batch_id1, 0, -1, rank_items1, sort_direction,)))
        threads[-1].start()
    if priority_batch_id2:
        threads.append(threading.Thread(target=get_redis_ranks_threaded,
                                        args=(REDIS_CLIENT, priority_batch_id2, 0, -1,  rank_items2, sort_direction,)))
        threads[-1].start()
    stage_details = {}
    department_details = {}
    threads.append(threading.Thread(target=get_stages_departments,
                                    args=(organization_id, stage_details, department_details,)))
    threads[-1].start()

    previous_rank_items = get_redis_previous_ranks(REDIS_CLIENT, priority_batch_id,
                                                   0,  -1, sort_direction)
    default_rank_items = get_redis_default_ranks(
        REDIS_CLIENT, priority_batch_id)

    no_rank_items = get_redis_no_ranks(REDIS_CLIENT, priority_batch_id,
                                       0, -1, sort_direction)
    rank_items = get_redis_ranks(REDIS_CLIENT, priority_batch_id,
                                 0, -1, sort_direction)

    print("rank_items", rank_items)
    print("no rank items", no_rank_items)
    filters = []

    if filter:
        if filter.get('byMe'):
            filters.append(
                {
                    "term": {
                        "userID.keyword": filter['byMe']
                    }
                }
            )
        if filter.get('Ranks'):
            # Check if one of the boolean values is False
            if not filter['Ranks']['Default']:
                selected_value = 'assigned'
            elif not filter['Ranks']['Assigned']:
                selected_value = 'default'
            else:
                selected_value = None

            if selected_value:
                if selected_value == 'default':
                    filters.append(
                        {
                            "range": {
                                "fixedRank": {"lte": -1}
                            }
                        }
                    )
                else:
                    filters.append(
                        {
                            "range": {
                                "fixedRank": {"gt": -1}
                            }
                        }
                    )

        if filter.get('Stage'):
            stages = [stage for stage, checked in filter['Stage'].items() if checked]
            
            filters.append(
                {
                    "terms": {
                        "levelID.keyword": stages
                    }
                }
            )
        if filter.get('Created'):
            filters.append(
                {
                    "range": {
                        "createdAt": {
                            "gte": filter['Created'][0],
                            "lte": filter['Created'][1]
                        }
                    }
                }
            )
        if filter.get('Department'):
            filters.append(
                {
                    "terms": {
                        "departments.keyword": filter['Department']
                    }
                }
            )

    if sort_field == "rank":

        if sort_direction == "ASC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            default_rank_items = {item.decode('utf-8'): int(rank)
                                  for item, rank in default_rank_items.items()}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items}
            print("dict of rank_items", rank_items)
            print("dict of no rank items", no_rank_items)
            all_items = list(rank_items.keys()) + list(no_rank_items.keys())
            print("all_items", all_items)

            print("all_items", all_items)

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": sort_direction.lower(),
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }
                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }

                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters
                }
            }

        elif sort_direction == "DESC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            default_rank_items = {item.decode('utf-8'): int(rank)
                                  for item, rank in default_rank_items.items()}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items[::-1]}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items[::-1]}
            all_items = list(no_rank_items.keys()) + list(rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": "asc",
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }


                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    elif sort_field == 'title':
        previous_rank_items = {item.decode('utf-8'): int(rank)
                               for item, rank in previous_rank_items}
        default_rank_items = {item.decode('utf-8'): int(rank)
                              for item, rank in default_rank_items.items()}
        rank_items = {item.decode('utf-8'): int(rank)
                      for item, rank in rank_items}
        no_rank_items = {item.decode('utf-8'): int(rank)
                         for item, rank in no_rank_items}
        all_items = list(rank_items.keys()) + list(no_rank_items.keys())

        if sort_direction == "ASC":

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [
                {"title.keyword": {"order": sort_direction.lower()}}]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

        elif sort_direction == "DESC":
            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [
                {"title.keyword": {"order": sort_direction.lower()}}]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    elif sort_field == 'level':
        previous_rank_items = {item.decode('utf-8'): int(rank)
                               for item, rank in previous_rank_items}
        default_rank_items = {item.decode('utf-8'): int(rank)
                              for item, rank in default_rank_items.items()}
        rank_items = {item.decode('utf-8'): int(rank)
                      for item, rank in rank_items[::-1]}
        no_rank_items = {item.decode('utf-8'): int(rank)
                         for item, rank in no_rank_items[::-1]}
        all_items = list(rank_items.keys()) + list(no_rank_items.keys())

        if sort_direction == "ASC":

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": sort_direction.lower(),
                    "script": {
                        "lang": "painless",
                        "source": "def levelID = doc['levelID.keyword'].value; if (params.scores.containsKey(levelID)) { return params.scores[levelID]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                stage: info['level'] for stage, info in stage_details.items()}
                        }

                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

        elif sort_direction == "DESC":

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": sort_direction.lower(),
                    "script": {
                        "lang": "painless",
                        "source": "def levelID = doc['levelID.keyword'].value; if (params.scores.containsKey(levelID)) { return params.scores[levelID]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                stage: info['level'] for stage, info in stage_details.items()}
                        }

                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    elif sort_field == 'change':
        rank_items = {item.decode('utf-8'): int(rank)
                      for item, rank in rank_items}

        if sort_direction == "ASC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            change_rank_items = {
                item: previous_rank_items.get(item, int(rank)) - int(rank)
                for item, rank in rank_items.items()
            }

            change_rank_items.update({item.decode('utf-8'): 0
                                      for item, rank in no_rank_items})
            change_rank_items = dict(
                sorted(change_rank_items.items(), key=lambda item: item[1]))
            all_items = list(change_rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": sort_direction.lower(),
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }


                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

        elif sort_direction == "DESC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            change_rank_items = {
                item: previous_rank_items.get(item, int(rank)) - int(rank)
                for item, rank in rank_items.items()
            }

            change_rank_items.update({item.decode('utf-8'): 0
                                      for item, rank in no_rank_items})
            change_rank_items = dict(
                sorted(change_rank_items.items(), key=lambda item: item[1], reverse=True))
            all_items = list(change_rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": "asc",
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }


                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    elif sort_field == 'previousRank':
        no_previous_rank_items = get_redis_no_previous_ranks(REDIS_CLIENT, priority_batch_id,
                                                             0,  -1, sort_direction)
        if sort_direction == "ASC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            no_previous_rank_items = {item.decode('utf-8'): int(rank)
                                      for item, rank in no_previous_rank_items}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items}
            all_items = list(previous_rank_items.keys()) + \
                list(no_previous_rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": sort_direction.lower(),
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }


                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

        elif sort_direction == "DESC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items[::-1]}
            no_previous_rank_items = {item.decode('utf-8'): int(rank)
                                      for item, rank in no_previous_rank_items[::-1]}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items}
            all_items = list(no_previous_rank_items.keys()) + \
                list(previous_rank_items.keys())
            # if filter.get('Department'):
            #     all_items = apply_department_filters(
            #         organization_id, all_items, departments_gammas, filter, query_threads)

            # all_items = all_items[start_rank:end_rank]

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "_script": {
                    "type": "number",
                    "order": "asc",
                    "script": {
                        "lang": "painless",
                        "source": "def id = doc['id.keyword'].value; if (params.scores.containsKey(id)) { return params.scores[id]; } else { return 1000; }",
                        "params": {
                            "scores": {
                                item: i for i, item in enumerate(all_items)
                            }
                        }


                    }
                }
            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    elif sort_field == 'createdAt':
        if sort_direction == "ASC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            default_rank_items = {item.decode('utf-8'): int(rank)
                                  for item, rank in default_rank_items.items()}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items[::-1]}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items[::-1]}
            all_items = list(rank_items.keys()) + list(no_rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "createdAt": {"order": sort_direction.lower()}

            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

        elif sort_direction == "DESC":
            previous_rank_items = {item.decode('utf-8'): int(rank)
                                   for item, rank in previous_rank_items}
            default_rank_items = {item.decode('utf-8'): int(rank)
                                  for item, rank in default_rank_items.items()}
            rank_items = {item.decode('utf-8'): int(rank)
                          for item, rank in rank_items[::-1]}
            no_rank_items = {item.decode('utf-8'): int(rank)
                             for item, rank in no_rank_items[::-1]}
            all_items = list(rank_items.keys()) + list(no_rank_items.keys())

            request = {"from": start, "size": PAGE_SIZE}
            request["sort"] = [{
                "createdAt": {"order": sort_direction.lower()}

            }]
            request["query"] = {
                "bool": {
                    "must": [
                        {
                            "term": {
                                "organizationID.keyword": organization_id
                            }
                        },
                        {
                            "wildcard": {
                                "title.keyword": {
                                    "value": "*" + search + "*" if search else "*",
                                    "case_insensitive": True
                                }
                            }
                        }
                    ],
                    "should": [
                        {
                            "terms": {
                                "id.keyword": all_items
                            }

                        }
                    ],
                    "minimum_should_match": 1,
                    "filter": filters

                }
            }

    request["_source"] = ["id", "levelID", "title", "friendlyId", "description",
                          "measurementFields", "userID", "fixedRank", "departments"]
    print("request", request)
    # chart = {
    #     "Top": 10,
    #     "Calender": ["2024-04-01", "2024-04-29"],
    #     "Interval": "week"
    # }
    if chart:

        request["query"]["bool"]["should"].append(
            {
                "nested": {
                    "path": "rankHistory",
                    "query": {
                        "terms": {
                            "rankHistory.D": generate_dates(chart['Calender'][0], chart['Calender'][1], chart['Interval'].lower())
                        }
                    },
                    "inner_hits": {
                    }
                }
            }
        )

    print("request", request)
    response = OPENSEARCH_CLIENT.search(
        body=request,
        index=INDEX_NAME
    )
    print("response", response)

    for thread in threads:
        thread.join()

    merged_rankings = calculate_merged_rankings(start, response, gamma_measurement_fields_aggregates, priority_batches, priority_batch_id, priority_batch_id1, priority_batch_id2,
                                                rank_items, rank_items1, rank_items2, previous_rank_items, default_rank_items, stage_details, department_details)

    if page != -1:
        payload["items"] = merged_rankings[(page-1)*20: (page)*20]
    else:
        payload["items"] = merged_rankings
    # payload["nextToken"] = json.dumps(
    #     gamma_measurement_fields_aggregates)
    payload["nextToken"] = "{}"
    # if chart == "true":
    #     for merged_ranking in payload["items"]:
    #         if s3_rank_history.get(merged_ranking["Gamma"]["id"]):
    #             merged_ranking.update(
    #                 {"Departments": [{"id": department, "name": department_details[department]} for department in merged_ranking["Gamma"]["departments"] if department_details.get(department)],
    #                     "Chart": {date: info for date, info in s3_rank_history[merged_ranking["Gamma"]["id"]].items() if date_range["start"] <= date <= date_range["end"]}})

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET'
        },
        'body': payload
    }


@app.route('/rankings/valid', cors=True, methods=['GET'], content_types=['application/json'])
def valid_override_rank():
    priorityBatchID = app.current_request.query_params["priorityBatchID"]
    organizationID = app.current_request.query_params["organizationID"]
    gammaID = app.current_request.query_params["id"]
    userID = app.current_request.query_params.get("userID", None)
    override_rank = int(app.current_request.query_params.get("rank", 0))
    authorizer = app.current_request.context.get("authorizer", {})
    claims = authorizer.get("claims", {})

    authorized = False
    if claims.get("SUPERADMIN"):
        authorized = True
    elif claims.get("organization"):
        if organizationID == claims['organization']:
            authorized = True

    if not authorized:
        raise UnauthorizedError("Not authorized")

    if not override_rank or not gammaID:
        raise BadRequestError("override rank not passed")

    redis_rankings = {}
    no_rankings = {}
    threads = []

    threads.append(threading.Thread(target=get_redis_rankings,
                                    args=(REDIS_CLIENT, priorityBatchID, redis_rankings, no_rankings, True, False, False, False)))
    threads[-1].start()
    fixed_rank_gammas = get_global_gamma_table().query(
        KeyConditionExpression=Key(
            'organizationID').eq(organizationID) & Key(
            'fixedRank').eq(override_rank),
        IndexName='byOrganization',
        ProjectionExpression='id'
    )
    print(fixed_rank_gammas)
    for thread in threads:
        thread.join()
    # Extracting ranks and finding the maximum
    offset = 0 if redis_rankings['items'].get(gammaID, None) else 1
    ranks = [value["rank"] for value in redis_rankings['items'].values()]
    if not ranks:
        max_rank = 1
    else:
        max_rank = max(ranks) + offset

    if 0 < override_rank <= max_rank and not fixed_rank_gammas['Items']:
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Headers': '*',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET'
            },
            'body': True
        }
    else:
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Headers': '*',
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET'
            },
            'body': False
        }


@app.route('/rankings/get', cors=True, methods=['GET'], content_types=['application/json'])
def get_ranking():
    gammaID = app.current_request.query_params.get("id", None)
    priorityBatchID = app.current_request.query_params.get(
        "priorityBatchID", None)
    threads = []
    rank = []
    user_details = {}
    department_details = {}
    authorizer = app.current_request.context.get("authorizer", {})
    claims = authorizer.get("claims", {})

    if gammaID and priorityBatchID:
        # implement authorization logic here
        # return UnauthorizedError("Not authorized")
        context = app.current_request.context
        print("context", context)
        # user_id = context['identity']['cognitoAuthenticationProvider'].split(
        #     ':')[-1]
        # if not user_id:
        #     raise UnauthorizedError("Not authorized")
        print("gammaID", gammaID)
        print("priorityBatchID", priorityBatchID)

        threads.append(threading.Thread(target=get_redis_rank,
                                        args=(REDIS_CLIENT, priorityBatchID, gammaID, rank,)))
        threads[-1].start()

        gamma = GQL_CLIENT.execute(get_gamma_query, variable_values={"id": gammaID})[
            "getGamma"]
        print("gamma", gamma)
        authorized = False
        if claims.get("SUPERADMIN"):
            authorized = True
        elif claims.get("organization"):
            if gamma["organizationID"] == claims['organization']:
                authorized = True

        if not authorized:
            raise UnauthorizedError("Not authorized")
        if gamma['contributors']:
            threads.append(threading.Thread(target=get_user_emails,
                                            args=(gamma['contributors'], user_details,)))
            threads[-1].start()
        if gamma['departments']:
            threads.append(threading.Thread(target=get_department_names,
                                            args=(gamma['departments'], department_details,)))
            threads[-1].start()

        layout = json.loads(gamma['level']['form']['layout'])
        print("json layout", layout)
        layout_lookup = {data["id"]: i for i, data in enumerate(layout)}
        print("json layout lookup", layout_lookup)

        # Create a list of keys to pop
        # keys_to_pop = []
        for thread in threads:
            thread.join()

        existing_departments = gamma['departments'][:]
        existing_contributors = gamma['contributors'][:]

        gamma['departments'] = {"items": [{"id": department, "name": department_details[department]}
                                          for department in existing_departments if department_details.get(department)]}
        gamma['contributors'] = {"items": [{**user_details[user]}
                                           for user in existing_contributors if user_details.get(user)]}

        for key, value in gamma.items():
            if key == 'level':
                gamma[key]['measurementFieldConnection']['items'] = [
                    item for item in gamma[key]['measurementFieldConnection']['items'] if item['measurementField']]

            if os.getenv(key):
                if layout_lookup.get(os.environ[key], None) is not None:
                    layout[layout_lookup[os.environ[key]]]["value"] = value
            #     keys_to_pop.append(key)
            # elif isinstance(value, dict):
            #     keys_to_pop.append(key)
        if gamma['additionalFields']:
            for field in gamma['additionalFields']:
                # if not os.getenv(field['name']):
                if layout_lookup.get(field['name'], None) is not None:
                    layout[layout_lookup[field['name']]
                           ]['value'] = field['value']
        if gamma['measurementFields']:
            for field in gamma['measurementFields']:
                if layout_lookup.get(field['measurementFieldID'], None) is not None:
                    layout[layout_lookup[field['measurementFieldID']]
                           ]['value'] = field['value']

        # updated_layout = [{'name': field['name'], 'value': field['value']} if layout_lookup.get(
        #     field['name']) is not None else field for field in gamma['additionalFields']]
        # Pop the keys from the dictionary
        # for key in keys_to_pop:
        #     gamma.pop(key)
        gamma.pop("additionalFields", [])
        gamma.pop("measurementFields", [])
        # for thread in threads:
        #     thread.join()
        gamma['rank'] = rank[0]
        # gamma['departments']['items'] = [{"id": department, "name": department_details[department]}
        #                                  for department in gamma["departments"] if department_details.get(department)]
        # gamma['contributors']['items'] = [{**user_details[user]}
        #                                   for user in gamma["contributors"] if user_details.get(user)]

        payload = {"layout": layout, **gamma}
    else:
        raise BadRequestError("gamma id or prioritybatch id not passed")
    # print("payload", payload)

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'GET'
        },
        'body': payload
    }


def add_created_updated_timestamps(inputs):
    timenow = generate_iso8601_string()

    for input in inputs:
        input['createdAt'] = timenow
        input['updatedAt'] = timenow
    return inputs


def add_typename(inputs, typename):
    for input in inputs:
        input['__typename'] = typename
    return inputs


def generate_iso8601_string():
    # Get the current UTC time
    now = datetime.utcnow()

    # Format the datetime object to ISO 8601 string
    iso8601_string = now.strftime('%Y-%m-%dT%H:%M:%S.%f')[:-3] + 'Z'

    return iso8601_string


@app.route('/rankings/update', cors=True, methods=['POST'], content_types=['application/json'])
def edit_ranking():
    event = app.current_request.to_dict()
    body = app.current_request.json_body
    print('received event:')
    print(event)
    print("body", body)
    layout = body['layout']
    dynamodb_client = get_global_dynamodb_resource()
    # Iterate through all the key-value pairs in the environment variables
    env_ids = {value: key for key, value in os.environ.items()
               if key.islower()}

    # implement authorization logic here
    # return UnauthorizedError("Not authorized")
    context = app.current_request.context
    print("context", context)
    authorizer = app.current_request.context.get("authorizer", {})
    claims = authorizer.get("claims", {})
    authorized = False
    sub = claims['sub']
    if claims.get("SUPERADMIN"):
        authorized = True
    elif claims.get("ADMIN"):
        if body["organizationID"] == claims['ADMIN']:
            authorized = True

    # check equality for sponsorId, userId or contributors before updating gamma
    if not authorized:
        res = DynamoDBHelper.get(get_global_dynamodb_resource(), GAMMA_TABLE_NAME, {"id": body['id']}, [
            'id', 'userID', 'sponsorID', 'contributors'])
        if sub in set([res.get('userID'), res.get('sponsorID'), *res.get('contributors', [])]):
            authorized = True
        if not authorized:
            raise UnauthorizedError("Not authorized")

    # function_mapping = {"departments": {1: create_department_gamma_relation,
    #                                     -1: delete_department_gamma_relation},
    #                     "contributors": {1: create_contributor_gamma_relation,
    #                                      -1: delete_contributor_gamma_relation}}
    layout_lookup = {data["id"]: i for i, data in enumerate(layout)}
    print("json layout lookup", layout_lookup)

    additionalFields = []
    measurementFields = []
    threads = []
    errors = []
    print("env ids", env_ids)
    try:
        for element in layout:
            if not env_ids.get(element['id']):
                if element.get("type", " ") == "measurement":
                    if element.get("newValue"):
                        measurementFields.append(
                            {"measurementFieldID": element['id'], "value": float(element["newValue"])})

                else:
                    if element.get("newValue"):
                        additionalFields.append({"name": element['id'], "value": [
                            {"text": val} for val in element["newValue"]]})
                # elif element.get("value"):
                #     additionalFields.append(
                #         {"name": element['id'], "value": element['value']})
            # else:
            #     table = function_mapping.get(env_ids[element['id']], None)
            #     if table:
            #         new_value = element.get('newValue', None)
            #         print("element id", element["id"])
            #         print("table", table)
            #         if isinstance(new_value, dict):
            #             for key, value in new_value.items():
            #                 # newValue = [{**val, "gammaId": id}
            #                 #             for val in value]
            #                 threads.append(threading.Thread(target=table[int(key)],
            #                                                 args=(dynamodb_client, value, errors,)))
            #                 threads[-1].start()
        if layout_lookup.get(os.environ['sponsor'], None) is not None:
            if layout[layout_lookup[os.environ['sponsor']]]["newValue"]:
                sponsorID = layout[layout_lookup[os.environ['sponsor']]
                                   ]["newValue"]["id"]
            elif layout[layout_lookup[os.environ['sponsor']]]["value"]:
                sponsorID = layout[layout_lookup[os.environ['sponsor']]
                                   ]["value"]["id"]
            else:
                sponsorID = None
        else:
            sponsorID = None

        if layout_lookup.get(os.environ['contributors'], None) is not None:
            if layout[layout_lookup[os.environ['contributors']]]["newValue"]:
                contributors = layout[layout_lookup[os.environ['contributors']]]["newValue"]
            # elif layout[layout_lookup[os.environ['contributors']]]["value"]:
            #     contributors = layout[layout_lookup[os.environ['contributors']]]["value"]
            else:
                contributors = []
        else:
            contributors = []

        if layout_lookup.get(os.environ['departments'], None) is not None:
            if layout[layout_lookup[os.environ['departments']]]["newValue"]:
                departments = layout[layout_lookup[os.environ['departments']]
                                     ]["newValue"]
            # elif layout[layout_lookup[os.environ['departments']]]["value"]:
            #     departments = layout[layout_lookup[os.environ['departments']]
            #                          ]["value"]
            else:
                departments = []
        else:
            departments = []

        input = {
            "id": body["id"],
            "title": layout[layout_lookup[os.environ['title']]]["newValue"],
            "description": layout[layout_lookup[os.environ['description']]]["newValue"],
            "sponsorID":  sponsorID,
            "additionalFields": additionalFields,
            "measurementFields": measurementFields,
            "updatedByIAM": False,
            "departments": departments,
            "contributors": contributors

        }
        params = {
            'input': input
        }
        print("modified input", input)
        update_gamma_response = GQL_CLIENT.execute(
            update_gamma_mutation, variable_values=params)
        print("update_gamma_response", update_gamma_response)
        update_rankings(LAMBDA_CLIENT, body["organizationID"], True)
        for thread in threads:
            thread.join()
        if errors:
            update_gamma_response["updateGamma"]["errors"] = errors

    except Exception as e:
        print("error in updating gamma", e)
        traceback.print_exc()
        raise ChaliceViewError()
    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'POST'
        },
        'body': update_gamma_response["updateGamma"]
    }


@app.route('/rankings/add', cors=True, methods=['POST'], content_types=['application/json'])
def add_ranking():
    event = app.current_request.to_dict()
    body = app.current_request.json_body
    print('received event:')
    print(event)
    print("body", body)
    # implement authorization logic here
    authorizer = app.current_request.context.get("authorizer", {})
    print("authorizer", authorizer)
    claims = authorizer.get("claims", {})
    authorized = False
    if claims.get("SUPERADMIN"):
        authorized = True
    elif claims.get("organization"):
        if body["organizationID"] == claims['organization']:
            authorized = True

    if not authorized:
        raise UnauthorizedError("Not authorized")
    layout = body['layout']
    dynamodb_client = get_global_dynamodb_resource()
    # Iterate through all the key-value pairs in the environment variables
    env_ids = {value: key for key, value in os.environ.items()
               if key.islower()}

    # implement authorization logic here
    # return UnauthorizedError("Not authorized")
    context = app.current_request.context
    print("context", context)
    # user_id = context['identity']['cognitoAuthenticationProvider'].split(
    #     ':')[-1]
    # if not user_id:
    #     raise UnauthorizedError("Not authorized")

    # function_mapping = {"departments": {1: create_department_gamma_relation,
    #                                     -1: delete_department_gamma_relation},
    #                     "contributors": {1: create_contributor_gamma_relation,
    #                                      -1: delete_contributor_gamma_relation}}
    layout_lookup = {data["id"]: i for i, data in enumerate(layout)}
    print("json layout lookup", layout_lookup)

    additionalFields = []
    measurementFields = []
    threads = []
    errors = []
    id = str(uuid.uuid4())
    try:
        for element in layout:
            if not env_ids.get(element['id']):
                if element.get("type", " ") == "measurement":
                    if element.get("value"):
                        measurementFields.append(
                            {"measurementFieldID": element['id'], "value": float(element["value"])})
                else:
                    if element.get("value"):
                        additionalFields.append({"name": element['id'], "value": [
                            {"text": val} for val in element["value"]]})
                # elif element.get("value"):
                #     additionalFields.append(
                #         {"name": element['id'], "value": element['value']})
            # else:
            #     table = function_mapping.get(env_ids[element['id']], None)
            #     if table:
            #         new_value = element.get('value', None)
            #         if isinstance(new_value, dict):
            #             for key, value in new_value.items():
            #                 newValue = [{**val, "gammaId": id}
            #                             for val in value]
            #                 threads.append(threading.Thread(target=table[int(key)],
            #                                                 args=(dynamodb_client, newValue, errors,)))
            #                 threads[-1].start()
        if layout_lookup.get(os.environ['sponsor'], None) is not None:
            if layout[layout_lookup[os.environ['sponsor']]]["value"]:
                sponsorID = layout[layout_lookup[os.environ['sponsor']]
                                   ]["value"]["id"]
            else:
                sponsorID = None
        else:
            sponsorID = None
        if layout_lookup.get(os.environ['contributors'], None) is not None:
            if layout[layout_lookup[os.environ['contributors']]]["value"]:
                contributors = layout[layout_lookup[os.environ['contributors']]]["value"]
            # elif layout[layout_lookup[os.environ['contributors']]]["value"]:
            #     contributors = layout[layout_lookup[os.environ['contributors']]]["value"]
            else:
                contributors = []
        else:
            contributors = []

        if layout_lookup.get(os.environ['departments'], None) is not None:
            if layout[layout_lookup[os.environ['departments']]]["value"]:
                departments = layout[layout_lookup[os.environ['departments']]
                                     ]["value"]
            # elif layout[layout_lookup[os.environ['departments']]]["value"]:
            #     departments = layout[layout_lookup[os.environ['departments']]
            #                          ]["value"]
            else:
                departments = []
        else:
            departments = []

        input = {
            "id": id,
            "friendlyId": int(GQL_CLIENT.execute(
                update_organization_gamma_count_mutation, variable_values={"id": body["organizationID"]})['updateOrganizationGammaCount']['gammaCount']),
            "title": layout[layout_lookup[os.environ['title']]]["value"],
            "description": layout[layout_lookup[os.environ['description']]].get("value", ""),
            "levelID": body['levelID'],
            "organizationID": body['organizationID'],
            "departments": departments,
            "contributors": contributors,
            "userID": body["userID"],
            "fixedRank": body.get('fixedRank', -1),
            "sponsorID":  sponsorID,
            "additionalFields": additionalFields,
            "measurementFields": measurementFields,
            "updatedByIAM": False

        }
        params = {
            'input': input
        }
        print("modified input", input)
        create_gamma_response = GQL_CLIENT.execute(
            create_gamma_mutation, variable_values=params)
        print("create_gamma_response", create_gamma_response)
        params = {
            "type": "ITEM_ADDED",
            "organization_id": body['organizationID'],
            "user_id": body["userID"],
            "gamma_id": id
        }
        response = LAMBDA_CLIENT.invoke(
            FunctionName=FUNCTION_SCHEDULEDCOMPARISONQUEUER_NAME,
            InvocationType='Event',
            Payload=json.dumps(params).encode('utf-8')
        )
        update_rankings(LAMBDA_CLIENT, body['organizationID'])

        if errors:
            create_gamma_response["createGamma"]["errors"] = errors

    except Exception as e:
        print("error in creating gamma", e)
        traceback.print_exc()
        raise ChaliceViewError()
    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'POST'
        },
        'body': create_gamma_response["createGamma"]
    }
