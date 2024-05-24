from io import BytesIO
from gql import gql
import AppSyncHelper
import os
import time
import traceback
import boto3
import json
import string
import random
import csv
import openpyxl
import re
from datetime import datetime, timedelta
from botocore.exceptions import ClientError
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from threading import Thread

USERPOOLID = os.environ['AUTH_STRALIGN5D0DB4AF_USERPOOLID']
FUNCTION_SAMPLEDATACREATOR_NAME = os.environ["FUNCTION_SAMPLEDATACREATOR_NAME"]
# SENDER_EMAIL = os.environ['SENDER_EMAIL']
SENDER_EMAIL = "noreply@stralign.com"
BUCKET_NAME = os.environ['STORAGE_STRALIGNUPLOADS_BUCKETNAME']
ENV = os.environ['ENV']
USER_IMPORT_COLUMN_COUNT = 7


cognito_client = boto3.client('cognito-idp')
ses_client = boto3.client('ses')
gql_client = AppSyncHelper.create_graphql_client()
lambda_client = boto3.client('lambda')

login_link = "https://app.stralign.com/login"
if ENV != 'live':
    login_link = os.environ['HOSTING_CLOUDFRONT_SECURE_URL'] + "/login"


def handler(event, context):
    print('received event:')
    print(event)
    event_type = event.get('type')

    if event_type == "invite_user":
        email = event.get('email')
        firstname = event.get('firstname')
        lastname = event.get('lastname')
        department = event.get('department')
        organization = event.get('organization')
        role = event.get('role')
        manager = event.get('manager')
        user_weight = event.get('user_weight')
        invite_user(email, firstname, lastname, department,
                    organization, role, manager, user_weight)
    if event_type == "signup_user":
        email = event.get('email')
        firstname = event.get('firstname')
        lastname = event.get('lastname')
        organization = event.get('organization')
        role = event.get('role')
        user_weight = event.get('user_weight')
        signup_user(email, firstname, lastname,
                    organization, role, user_weight)

    elif event_type == "bulk_invite_users":
        file_key = event['file_key']
        sender_id = event['inviter']
        bulk_invite(file_key, sender_id)

    elif event_type == "resend_invite":
        userID = event['userID']
        resend_invite(userID)
    elif event_type == "reset_password":
        userID = event['userID']
        send_reset_password_mail(userID)
        sign_out_user(userID)

    elif event_type == "bulk_resend_invite":
        userIDs = event['userIDs']
        userIDs = userIDs.split(",")
        for userID in userIDs:
            resend_invite(userID)

    elif event_type == "bulk_send_missyou":
        users = event['EmailIDs']
        for user in users:
            send_email_missyou(user)

    return {
        'statusCode': 200,
        'headers': {
            'Access-Control-Allow-Headers': '*',
            'Access-Control-Allow-Origin': '*',
            'Access-Control-Allow-Methods': 'OPTIONS,POST,GET'
        },
        'body': json.dumps('Hello from your new Amplify Python lambda!')
    }


def get_graphql_client():
    print("creating a new graphql client")
    gql_client = AppSyncHelper.create_graphql_client()
    return gql_client


def sign_out_user(user_id):
    query = gql(
        """
            mutation createUserSignOutStatus($userID: ID!, ) {
                createUserSignOutStatus(input: {userID: $userID}) {
                    id
                    userID
                }
                    
                }
        """
    )
    params = {
        "userID": user_id
    }
    gql_response = gql_client.execute(query, variable_values=params)


def create_organization(organization_name):
    query = gql(
        """
            mutation CreateOrganization($name: String!, ) {
                createOrganization(input: {name: $name}) {
                    id
                    name
                }
                    
                }
        """
    )
    params = {
        "name": organization_name
    }
    gql_response = gql_client.execute(query, variable_values=params)

    response = cognito_client.create_group(
        GroupName=gql_response["createOrganization"]["id"],
        UserPoolId=os.environ['AUTH_STRALIGN5D0DB4AF_USERPOOLID'],
        Description=organization_name
    )
    print(response)

    return gql_response


def send_email_attachment(destination, successes, failed, subject, username, filename):
    # emailData = {
    #     "email": "email",
    #     "password": "password",
    #     "login_link": "https://d25d5goqjgy5td.cloudfront.net/login"
    # }
    emailData = {
        "total": successes + failed,
        "successes": successes,
        "failures": failed,
        "username": username,
        "filename": filename
    }
    now = datetime.now()
    est = now - timedelta(hours=4)
    try:
        msg = MIMEMultipart()
        msg['From'] = SENDER_EMAIL
        msg['To'] = destination

        msg['Subject'] = est.strftime('%m-%d-%Y') + " " + subject
        body = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8" />
            <meta http-equiv="X-UA-Compatible" content="IE=edge" />
            <meta name="viewport" content="width=device-width, initial-scale=1.0" />
            <title> StrAlign</title>

            <style>
            body,
            html {{
                font-family: Arial, Helvetica, sans-serif;
            }}
            p {{
                font-size: 15px;
                line-height: 22px;
                color: #161616;
            }}
            </style>
        </head>
        <body topmargin="0">
            <table
            width="600"
            border="0"
            align="center"
            cellpadding="0"
            cellspacing="0"
            class="bgtext"
            >
            <thead>
                <tr>
                <th width="100%;" style="display: block; margin: 20px 0">
                    <a href="#"
                    ><img
                        class="header-logo-image"
                        src="https://stralign-static-files.s3.amazonaws.com/logo.png"
                        alt="StrAlign"
                        title="StrAlign"
                    /></a>
                </th>
                </tr>
            </thead>

            <tbody style="background-color: #f8f8f8">
                <tr>
                <td colspan="1">
                    <table
                    width="100%"
                    border="0"
                    align="center"
                    cellpadding="8"
                    cellspacing="0"
                    class="bgtext"
                    >
                    <tr>
                        <div
                        style="
                            background-color: white;
                            border-radius: 12px;
                            width: 85%;
                            margin: 20px auto 20px;
                            padding: 5px 20px 30px;
                        "
                        >
                        <h4><b>Dear {username},</b></h4>
                        <p>Please find the report for your latest import.</p>
                        <p>Import file: <b>{filename}</b></p>
                        <p>
                            Total Imported Records:
                            <span style="color: #0070c0; font-weight: 600">{total}</span>
                        </p>
                        <div style="display: flex; margin-top: 20px">
                            <div
                            style="
                                background-color: #15ae1b1a;
                                padding: 5px 20px;
                                border-radius: 10px;
                                margin-right: 10px;
                            "
                            >
                            <p style="margin-bottom: 0">Successful Records</p>
                            <h3
                                style="color: #92d050; font-weight: 600; margin-top: 0"
                            >
                                {successes}
                            </h3>
                            </div>
                            <div
                            style="
                                background-color: #f932321a;
                                padding: 5px 20px;
                                border-radius: 10px;
                                margin-right: 10px;
                            "
                            >
                            <p style="margin-bottom: 0">Failed Records</p>
                            <h3
                                style="color: #f93232; font-weight: 600; margin-top: 0"
                            >
                                {failures}
                            </h3>
                            </div>
                        </div>
                        </div>
                    </tr>

                    <tr>
                        <th
                        width="100%;"
                        style="display: block; margin: 20px 0 0; padding-bottom: 0"
                        >
                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/Facebook.png"
                            alt="fb"
                            style="height: 21px"
                        /></a>
                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/TwitterCircle.png"
                            alt="insta"
                            style="height: 21px"
                        /></a>

                        <a href="#"
                            ><img
                            src="https://stralign-static-files.s3.amazonaws.com/Linkedin.png"
                            alt="linkedin"
                            style="height: 21px"
                        /></a>
                        </th>
                        <th
                        width="100%;"
                        style="display: block; margin: 0; padding-top: 0"
                        >
                        <p
                            style="
                            margin-top: 0;
                            color: #646464;
                            font-weight: 500;
                            font-size: 13px;
                            "
                        >
                            © 2024 Echo Consulting, LLC<br />
                            Copyright all rights reserved.
                        </p>
                        </th>
                    </tr>
                    </table>
                </td>
                </tr>
            </tbody>
            <tfoot>
                <tr>
                <td></td>
                <td></td>
                </tr>
            </tfoot>
            </table>
        </body>
        </html>

        """.format(**emailData)
        msg.attach(MIMEText(body, 'html'))

        # add the attachments
        attachments = ['/tmp/success.csv', '/tmp/failed.csv']
        for attachment in attachments:
            part = MIMEBase('application', "octet-stream")
            with open(attachment, 'rb') as file:
                part.set_payload(file.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition', f'attachment; filename="{os.path.basename(attachment)}"')
            msg.attach(part)

        # send the email
        raw_msg = msg.as_string()
        response = ses_client.send_raw_email(
            Source=SENDER_EMAIL,
            Destinations=[destination],
            RawMessage={'Data': raw_msg}
        )
        print("Email sent! Message ID:", response['MessageId'])

    except ClientError as e:
        print("Email not sent. Error message: ",
              e.response['Error']['Message'])

    return {
        'statusCode': 200
    }


def generate_csv(data, filename):
    with open(filename, mode='w', newline='') as file:
        writer = csv.writer(file)
        for row in data:
            writer.writerow(row)


def get_organization_departments(organization_id):
    departments = []
    nextToken = None
    count = 1
    while (nextToken != None or count < 2):
        count += 1
        query = gql(
            """
            query MyQuery($organizationID: ID!, $nextToken: String) {
                getOrganization(id: $organizationID) {
                    Departments(nextToken: $nextToken) {
                    items {
                        id
                        name
                    }
                    nextToken
                    }
                }
            }

            """
        )
        params = {
            "organizationID": organization_id
        }
        response = gql_client.execute(query, variable_values=params)
        print(response)
        nextToken = response['getOrganization']['Departments']["nextToken"]
        departments.extend(response['getOrganization']['Departments']["items"])

    print("--> ", departments)
    departments_dictionary = {item['name']: item['id'] for item in departments}
    return departments_dictionary

# Invite new user


def signup_user(email, firstname, lastname, organization_name, role, user_weight, gql_client=gql_client):
    ''' 
        Create user in cognito with temporary password 
        Create organization through /signup
        Add user to role and org cognito groups
        Send email invite with temporary password
        Create user in database using graphQL query
    '''
    try:
        role = str(role).upper()
        role = "ADMIN" if role == "SUPERADMIN" else role
        characters = string.ascii_letters + string.digits
        password = ''.join(random.choice(characters) for i in range(24))
        cognito_user_created = cognito_client.admin_create_user(
            UserPoolId=USERPOOLID,
            Username=email,
            UserAttributes=[
                {
                    'Name': 'email',
                    'Value': email
                },
                {
                    'Name': 'family_name',
                    'Value': lastname
                },
                {
                    'Name': 'given_name',
                    'Value': firstname
                },
                {
                    'Name': 'email_verified',
                    'Value': 'true'
                }
            ],
            TemporaryPassword=password,
            MessageAction='SUPPRESS',
            DesiredDeliveryMediums=[]
        )
        print("user created :: ", cognito_user_created)
        userID = cognito_user_created['User']['Username']

        create_organization_response = create_organization(organization_name)
        organization = create_organization_response["createOrganization"]["id"]
        # add_user_to_group(userID, organization)
        add_user_to_group(userID, str(role).upper())
        # create_user_gql(userID, email, firstname, lastname,
        #                 organization, role, user_weight, gql_client)
        create_user_gql(userID, email, firstname, lastname,
                        None, None, organization, role, user_weight, gql_client)
        payload = {
            "org_id": organization,
            "org_name": organization_name,
            "user_id": userID
        }
        response = lambda_client.invoke(
            FunctionName=FUNCTION_SAMPLEDATACREATOR_NAME,
            InvocationType='Event',
            Payload=json.dumps(payload).encode('utf-8'),
        )
        username = firstname
        send_email_invite(email, password, username, 'inviteUser')
        return userID

    except Exception as e:
        print(e)
        if isinstance(e, ClientError):
            error_code = e.response['Error'].get('Code')
            if error_code == 'UsernameExistsException':
                raise Exception(f"Exception !! {email} already exists")
            else:
                raise Exception("An error occurred:", e)
        else:
            raise Exception("Something went wrong")


def invite_user(email, firstname, lastname, department, organization, role, manager, user_weight, gql_client=gql_client):
    ''' 
        Create user in cognito with temporary password 
        Add user to role and org cognito groups
        Send email invite with temporary password
        Create user in database using graphQL query
    '''
    try:
        role = str(role).upper()
        role = "ADMIN" if role == "SUPERADMIN" else role
        characters = string.ascii_letters + string.digits
        password = ''.join(random.choice(characters) for i in range(24))
        cognito_user_created = cognito_client.admin_create_user(
            UserPoolId=USERPOOLID,
            Username=email,
            UserAttributes=[
                {
                    'Name': 'email',
                    'Value': email
                },
                {
                    'Name': 'family_name',
                    'Value': lastname
                },
                {
                    'Name': 'given_name',
                    'Value': firstname
                },
                {
                    'Name': 'email_verified',
                    'Value': 'true'
                }
            ],
            TemporaryPassword=password,
            MessageAction='SUPPRESS',
            DesiredDeliveryMediums=[]
        )
        print("user created :: ", cognito_user_created)
        userID = cognito_user_created['User']['Username']

        if user_weight is None or user_weight == '':
            try:
                organization_weights = get_user_weights_for_organization(
                    organization, gql_client)
                print("organization weights --- ", organization_weights)
                for weightSetting in organization_weights:
                    if weightSetting['role'] == role:
                        user_weight = weightSetting['weight']
            except Exception as e:
                traceback.print_exc()
                user_weight = 1

        # add_user_to_group(userID, organization)
        add_user_to_group(userID, str(role).upper())
        create_user_gql(userID, email, firstname, lastname,
                        department, manager, organization, role, user_weight, gql_client)
        username = firstname
        send_email_invite(email, password, username, 'inviteUser')
        return userID

    except Exception as e:
        print(e)
        if isinstance(e, ClientError):
            error_code = e.response['Error'].get('Code')
            if error_code == 'UsernameExistsException':
                raise Exception(f"Exception !! {email} already exists")
            else:
                raise Exception("An error occurred:", e)
        else:
            raise Exception("Something went wrong")


def invite_worker(chunk, organizationID, successfull_records, unsuccessfull_records):
    my_gql_client = get_graphql_client()
    for user in chunk:
        try:
            email = user['email']
            firstname = user['first name']
            lastname = user['last name']
            department = user['departmentID']
            organization = organizationID
            role = user['role']
            manager = user['managerID']
            weight = user['weight']
            invite_user(email, firstname, lastname, department,
                        organization, role, manager, weight, my_gql_client)
            user.pop('departmentID', None)
            user.pop('managerID', None)
            successfull_records.append(list(user.values()))
        except Exception as E:
            traceback.print_exc()
            print(user)
            user.pop('departmentID', None)
            user.pop('managerID', None)
            unsuccessfull_records.append(list(user.values()))


def bulk_invite(file_key, sender_id):
    print(file_key)
    user_details = get_user_details(sender_id)
    destination = user_details["email"]
    organizationID = user_details['organizationID']
    print("Organization ID ::", organizationID)
    file_data = extract_data_from_file(file_key)
    all_records = file_data['all_data']
    column_names = file_data['columns']
    managers_departments_data = get_managers_departments(all_records)
    print("Managers and Departments :: ", managers_departments_data)
    managers_email_list = list(managers_departments_data['manager'])
    print("manager emails :: ", managers_email_list)
    managers_dictionary = get_managers(
        managers_email_list)
    department_dictionary = get_organization_departments(
        organizationID)

    try:
        organization_weights = get_user_weights_for_organization(
            organizationID)
        print("organization weights --- ", organization_weights)
        if organization_weights != None:
            converted_weights = {item["role"].upper(): item["weight"]
                                 for item in organization_weights}
        else:
            converted_weights = None
    except Exception as e:
        traceback.print_exc()

    for record in all_records:
        try:
            record['email'] = record['email'].lower()
            record['departmentID'] = department_dictionary.get(
                record['department'], None)
            if not record['departmentID']:
                record['department'] = ''
            record['managerID'] = managers_dictionary.get(
                record['manager'], None)
            if not record['managerID']:
                record['manager'] = ''
            if "weight" in record and (record['weight'] is None or record['weight'] == ''):
                record['weight'] = converted_weights.get(
                    record['role'].upper()) if converted_weights is not None else 1
        except Exception as e:
            # print("-- KEY ERROR FOR EITHER DEPARTMENT OR MANAGER --")
            print(record)
            traceback.print_exc()

    unsuccessfull_records = []
    successfull_records = []
    NUMBER_OF_THREADS = 4
    chunk_size = len(all_records) // NUMBER_OF_THREADS
    if chunk_size >= 1:
        user_chunks = [all_records[i:i+chunk_size]
                       for i in range(0, len(all_records), chunk_size)]
    else:
        user_chunks = [all_records]
    print("chunk size", chunk_size)
    print(user_chunks)
    threads = []
    for chunk in user_chunks:
        try:
            thread = Thread(target=invite_worker, args=(
                chunk, organizationID, successfull_records, unsuccessfull_records,))
            thread.start()
            threads.append(thread)
        except Exception as E:
            traceback.print_exc()

    # WAIT FOR ALL THREAD TO COMPLETE EXECUTION
    for thread in threads:
        thread.join()

    unsuccessfull_records.insert(0, column_names)
    successfull_records.insert(0, column_names)
    print("UNSUCCESSFUL RECORDS - ", unsuccessfull_records)
    print("SUCCESSFUL RECORDS ", successfull_records)
    generate_csv(successfull_records, '/tmp/success.csv')
    generate_csv(unsuccessfull_records, '/tmp/failed.csv')
    send_email_attachment(destination, len(successfull_records)-1,
                          len(unsuccessfull_records)-1, "Bulk User Import Summary", user_details["firstName"], file_key.split('/')[-1])


def resend_invite(userID):
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for i in range(24))

    response = cognito_client.admin_set_user_password(
        UserPoolId=USERPOOLID,
        Username=userID,
        Password=password,
        Permanent=False
    )
    print(response)
    user_details = get_user_details(userID)
    user_email = user_details['email']
    username = user_details['firstName']
    send_email_invite(user_email, password, username, 'remindUser')


def send_reset_password_mail(userID):
    characters = string.ascii_letters + string.digits
    password = ''.join(random.choice(characters) for i in range(24))

    response = cognito_client.admin_set_user_password(
        UserPoolId=USERPOOLID,
        Username=userID,
        Password=password,
        Permanent=False
    )
    print(response)
    user_details = get_user_details(userID)
    user_email = user_details['email']
    username = user_details['firstName']
    send_email_invite(user_email, password, username, 'resetPassword')


def send_email_invite(email, password, username, template_name,  **userData):
    '''Send email with temporary password'''

    emailData = {
        "email": email,
        "password": password,
        "username": username,
        "login_link": login_link
    }

    print("User Data", userData)
    sendEmail = ses_client.send_templated_email(
        Source=SENDER_EMAIL,
        Destination={
            'ToAddresses': [email]
        },
        Template=template_name,
        TemplateData=json.dumps(emailData))
    print(sendEmail)
    return


def send_email_missyou(user):
    '''Send Miss you emails to users'''

    print("User data", user)
    emailData = {
        "username": user[1],
        "login_link": login_link
    }
    sendEmail = ses_client.send_templated_email(
        Source=SENDER_EMAIL,
        Destination={
            'ToAddresses': [user[0]]
        },
        Template='missYou',
        TemplateData=json.dumps(emailData))
    print(sendEmail)
    return


def add_user_to_group(userID, group_name):
    ''' Add user to cognito groups '''
    response = cognito_client.admin_add_user_to_group(
        UserPoolId=USERPOOLID,
        Username=userID,
        GroupName=group_name
    )
    print("Added user to group :: ", response)
    return userID


def get_managers_departments(all_data):
    data = {"manager": set(), "department": set()}
    for record in all_data:
        data['manager'].add(record['manager'])
        data['department'].add(record['department'])
    return data


def remove_escape_characters(item):
    try:
        escape_pattern = re.compile(r'[\n\t\r]')
        translation_table = str.maketrans('', '', '\n\t\r')
        return escape_pattern.sub('', item).translate(translation_table)
    except:
        return item


def extract_data_from_file(file_key):
    s3 = boto3.client('s3')
    response = s3.get_object(Bucket=BUCKET_NAME, Key=file_key)

    data = response['Body'].read()
    # Parse CSV data
    all_data = []
    if file_key.endswith('.csv'):
        print("Reading CSV -- {:.4f}".format(time.time()))
        data = data.decode('utf-8').splitlines()
        csv_reader = csv.reader(data)

        # column names from first row of CSV data
        column_names = next(csv_reader)
        column_names = [col.lower()
                        for col in column_names if col is not None and col.strip()]
        print(column_names)

        # remaining rows of CSV data
        for row in csv_reader:
            row = row[:USER_IMPORT_COLUMN_COUNT]
            if len(set(row)) != 1:
                user_data = list(map(remove_escape_characters, row))
                all_data.append(dict(zip(column_names, user_data)))

        print("Reading CSV completed -- {:.4f}".format(time.time()))

    elif file_key.endswith('.xlsx'):
        start = time.time()
        print("Reading XLSX file -- {:.4f}".format(start))
        with BytesIO(data) as f:
            book = openpyxl.load_workbook(filename=f, read_only=True)
            sheet = book.active
            # column_names = next(sheet.iter_rows(
            #     min_row=1, max_row=1, values_only=True))
            # column_names = [col.lower() for col in column_names]
            column_names = [sheet.cell(row=1, column=col).value.lower() for col in range(1, USER_IMPORT_COLUMN_COUNT + 1) if
                            sheet.cell(row=1, column=col).value]
            # Read data from sheet
            count = 0
            for row in sheet.iter_rows(min_row=2, values_only=True):  # SKIP HEADERS
                row = row[:USER_IMPORT_COLUMN_COUNT]
                count += 1
                if len(set(row)) != 1:
                    user_data = list(map(remove_escape_characters, row))
                    all_data.append(dict(zip(column_names, user_data)))
            print(count)
            print("Reading completed -- {:.4f}".format(time.time() - start))
    return {'all_data': all_data, 'columns': column_names}


def get_user_weights_for_organization(organization_id, gql_client=gql_client):
    organization_details = get_organization_gql(organization_id, gql_client)
    user_weights = organization_details['weightSettings']
    print(user_weights)
    return user_weights


def get_organization_gql(organization_id, gql_client=gql_client):
    query = gql(
        """
        query MyQuery($organizationID: ID!) {
            getOrganization(id: $organizationID) {
                id
                name
                weightSettings {
                    role
                    weight
                }
            }
        }
        """
    )
    params = {
        "organizationID": organization_id
    }
    response = gql_client.execute(query, variable_values=params)
    print(response)
    return response['getOrganization']


def get_user_details(userID):
    query = gql(
        """
        query MyQuery($userID: ID!) {
            getUser(id: $userID) {
                email
                firstName
                organizationID
            }
        }
        """
    )
    params = {
        "userID": userID
    }
    response = gql_client.execute(query, variable_values=params)
    print(response)
    return response['getUser']


def create_user_gql(id, email, firstname, lastname, department, manager, organization, role, weight, gql_client=gql_client):
    '''Create user in database using graphQL query'''
    query = gql(
        """
        mutation MyMutation($role: USER_ROLE, $organizationID: ID, $lastName: String, $firstName: String, $id: ID, $email: AWSEmail!, $departmentID: ID, $managerID: ID, $weight: Float) {
            createUser(input: {email: $email, firstName: $firstName, lastName: $lastName, organizationID: $organizationID, role: $role, id: $id, departmentID: $departmentID, managerID: $managerID, weight: $weight, status: INVITED}) {
                id
            }
        }
        """
    )
    params = {
        "email": email,
        "firstName": firstname,
        "lastName": lastname,
        "organizationID": organization,
        "role": role,
        "id": id,
        "weight": weight
    }
    if manager:
        params['managerID'] = manager
    if department:
        params['departmentID'] = department

    print(params)
    response = gql_client.execute(query, variable_values=params)
    print(response)


def get_managers(manager_email_list):
    count = len(manager_email_list)
    filter_str = [f"$email{i}" for i in range(1, count+1)]
    input_list = [f"$email{i}:String" for i in range(1, count+1)]

    print("----", filter_str)
    print("====", input_list)
    # Create the list of dictionaries
    filter_list = []
    for email in filter_str:
        filter_list.append({"email": {"eq": email}})
    filter_str = ','.join(str(i) for i in filter_list).replace("'", "")
    input_str = ','.join(str(i) for i in input_list).replace("'", "")

    print("----", filter_str)
    print("====", input_str)

    query = gql(
        f"""
        query MyQuery({input_str}) {{
            listUsers(filter: {{or:[{filter_str}]}}) {{
                items {{
                    email
                    id
                }}
            }}
        }}
        """
    )
    params = {}
    for i in range(1, count+1):
        params[f"email{i}"] = manager_email_list[i - 1]
    print("params ::", params)
    response = gql_client.execute(query, variable_values=params)
    response = response['listUsers']['items']
    print(response)
    managers_dictionary = {item['email']: item['id'] for item in response}
    return managers_dictionary
